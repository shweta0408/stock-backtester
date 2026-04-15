"""
Stock Backtesting & Analysis System  (OPTIMISED)
=================================================
Speed improvements over original:
  • Batch yfinance download  → all tickers in ONE API call
  • ThreadPoolExecutor       → parallel per-stock processing
  • Reduced NSE timeout      → 5 s instead of 12 s, optional
  • Trimmed sleep/retry delays
  • Pre-warmed in-memory cache before main loop

Usage:
    python backtest.py screener_input.xlsx final_output.xlsx
    python backtest.py screener_input.csv  final_output.xlsx

Requirements:
    pip install yfinance pandas numpy openpyxl requests
"""

import os
import time
import logging
import warnings
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
from typing import Optional

import pandas as pd
import numpy as np
import requests

try:
    import yfinance as yf
    YF_AVAILABLE = True
except ImportError:
    YF_AVAILABLE = False
    print("WARNING: yfinance not installed. Run: pip install yfinance")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ─── Logging Setup ─────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("StockBacktest")

# ── Parallelism config ─────────────────────────────────────────────────────────
MAX_WORKERS = 10    # concurrent threads for per-stock processing
BATCH_CHUNK_SIZE = 50    # max tickers per yf.download() batch call
FETCH_DELIVERY = True  # set False to skip NSE delivery (saves ~1-2 s/stock)
NSE_TIMEOUT = 5     # seconds per NSE request


# ═══════════════════════════════════════════════════════════════════════════════
#  1. DATA LOADING
# ═══════════════════════════════════════════════════════════════════════════════

def load_input_file(filepath: str) -> pd.DataFrame:
    """Load Excel or CSV screener file and normalise column names."""
    ext = os.path.splitext(filepath)[1].lower()
    if ext in (".xlsx", ".xls"):
        df = pd.read_excel(filepath)
    elif ext == ".csv":
        df = pd.read_csv(filepath)
    else:
        raise ValueError(f"Unsupported file type: {ext}")

    col_map = {}
    for col in df.columns:
        lc = col.strip().lower()
        if lc in ("symbol", "ticker", "scrip", "stock"):
            col_map[col] = "Symbol"
        elif lc in ("sector", "industry"):
            col_map[col] = "Sector"
        elif lc in ("market cap", "marketcap", "mkt cap", "market_cap"):
            col_map[col] = "Market Cap"
        elif "date" in lc or lc in ("entry", "entry date", "date of entry"):
            col_map[col] = "Date of Entry"

    df.rename(columns=col_map, inplace=True)

    for required in ("Symbol", "Date of Entry"):
        if required not in df.columns:
            raise ValueError(
                f"Required column '{required}' not found. "
                f"Available columns: {list(df.columns)}"
            )

    df["Symbol"] = df["Symbol"].astype(str).str.strip().str.upper()
    df["Date of Entry"] = pd.to_datetime(df["Date of Entry"], dayfirst=True)

    if "Sector" not in df.columns:
        df["Sector"] = "N/A"
    if "Market Cap" not in df.columns:
        df["Market Cap"] = "N/A"

    log.info(f"Loaded {len(df)} rows from '{filepath}'")
    return df


# ═══════════════════════════════════════════════════════════════════════════════
#  2. HISTORICAL DATA — batch pre-fetch + in-memory cache
# ═══════════════════════════════════════════════════════════════════════════════

_DATA_CACHE: dict = {}


def _batch_prefetch(symbols: list[str], start: datetime, end: datetime,
                    exchange: str = "NS"):
    """
    Download all tickers in one yf.download() call (much faster than N calls).
    Populates _DATA_CACHE so per-stock lookups are instant.
    """
    if not YF_AVAILABLE or not symbols:
        return

    tickers = [f"{s}.{exchange}" for s in symbols]
    fetch_start = (start - timedelta(days=30)).strftime("%Y-%m-%d")
    fetch_end = (end + timedelta(days=10)).strftime("%Y-%m-%d")

    log.info(f"Batch downloading {len(tickers)} tickers ({exchange}) …")
    try:
        raw = yf.download(
            tickers,
            start=fetch_start,
            end=fetch_end,
            auto_adjust=True,
            progress=False,
            timeout=60,
            group_by="ticker",
        )
    except Exception as e:
        log.warning(f"Batch download failed: {e}")
        return

    if raw is None or raw.empty:
        return

    # yf.download with group_by="ticker" returns MultiIndex columns: (ticker, OHLCV)
    if isinstance(raw.columns, pd.MultiIndex):
        available_tickers = raw.columns.get_level_values(0).unique()
        for t in available_tickers:
            symbol = t.split(".")[0]
            try:
                df_t = raw[t].copy()
                df_t.index = pd.to_datetime(df_t.index).tz_localize(None)
                df_t = df_t.dropna(how="all").sort_index()
                if not df_t.empty:
                    cache_key = f"{t}_{fetch_start}_{fetch_end}"
                    _DATA_CACHE[cache_key] = df_t
            except Exception:
                pass
    else:
        # Single ticker — treat as individual
        if len(tickers) == 1:
            t = tickers[0]
            symbol = symbols[0]
            raw.index = pd.to_datetime(raw.index).tz_localize(None)
            raw = raw.sort_index()
            cache_key = f"{t}_{fetch_start}_{fetch_end}"
            _DATA_CACHE[cache_key] = raw

    log.info(f"Batch cache populated: {len(_DATA_CACHE)} entries")


def fetch_data(symbol: str, start: datetime, end: datetime,
               exchange: str = "NS") -> Optional[pd.DataFrame]:
    """
    Fetch OHLCV — returns from cache if available, else falls back to
    individual download (also tries .BO if .NS empty).
    """
    if not YF_AVAILABLE:
        return None

    fetch_start = (start - timedelta(days=30)).strftime("%Y-%m-%d")
    fetch_end = (end + timedelta(days=10)).strftime("%Y-%m-%d")

    for suffix in (exchange, "BO"):
        ticker = f"{symbol}.{suffix}"
        cache_key = f"{ticker}_{fetch_start}_{fetch_end}"

        if cache_key in _DATA_CACHE:
            cached = _DATA_CACHE[cache_key]
            if cached is not None and not cached.empty:
                return cached
            continue   # cached as empty → try next exchange

        # Individual fallback (not pre-cached)
        for attempt in range(3):
            try:
                raw = yf.download(
                    ticker,
                    start=fetch_start,
                    end=fetch_end,
                    auto_adjust=True,
                    progress=False,
                    timeout=25,
                )
                if raw is not None and not raw.empty:
                    if isinstance(raw.columns, pd.MultiIndex):
                        raw.columns = raw.columns.get_level_values(0)
                    raw.index = pd.to_datetime(raw.index).tz_localize(None)
                    raw = raw.sort_index()
                    _DATA_CACHE[cache_key] = raw
                    return raw
                else:
                    _DATA_CACHE[cache_key] = None
                    break
            except Exception as e:
                log.warning(
                    f"[{symbol}] Attempt {attempt+1}/3 ({ticker}): {e}")
                time.sleep(1.0 * (attempt + 1))   # shorter back-off

    log.error(f"[{symbol}] All fetch attempts failed.")
    return None


def get_nearest_trading_day_idx(df: pd.DataFrame,
                                target_date: datetime) -> Optional[int]:
    target = pd.Timestamp(target_date).normalize()
    mask = df.index <= target
    if not mask.any():
        return None
    return int(np.where(mask)[0][-1])


# ═══════════════════════════════════════════════════════════════════════════════
#  3. RSI
# ═══════════════════════════════════════════════════════════════════════════════

def calculate_rsi(prices: pd.Series, period: int = 14) -> pd.Series:
    delta = prices.diff()
    gain = delta.clip(lower=0)
    loss = (-delta).clip(lower=0)
    alpha = 1.0 / period
    avg_gain = gain.ewm(alpha=alpha, min_periods=period, adjust=False).mean()
    avg_loss = loss.ewm(alpha=alpha, min_periods=period, adjust=False).mean()
    rs = avg_gain / avg_loss.replace(0.0, np.nan)
    rsi = 100.0 - (100.0 / (1.0 + rs))
    rsi = rsi.where(avg_loss != 0, 100.0)
    return rsi


# ═══════════════════════════════════════════════════════════════════════════════
#  4. FORWARD RETURNS
# ═══════════════════════════════════════════════════════════════════════════════

def compute_returns(df: pd.DataFrame, entry_idx: int,
                    windows: tuple = (5, 10, 15)) -> dict:
    closes = df["Close"]
    entry_price = float(closes.iloc[entry_idx])
    result = {"entry_price": round(entry_price, 2)}
    for w in windows:
        fwd_idx = entry_idx + w
        if fwd_idx < len(closes):
            fwd_price = float(closes.iloc[fwd_idx])
            pct = ((fwd_price - entry_price) / entry_price) * 100.0
            result[f"ret_{w}d"] = round(pct, 2)
        else:
            result[f"ret_{w}d"] = None
    return result


# ═══════════════════════════════════════════════════════════════════════════════
#  5. NSE DELIVERY  (optional, shorter timeout)
# ═══════════════════════════════════════════════════════════════════════════════

_NSE_SESSION: Optional[requests.Session] = None
_NSE_HEADERS = {
    "User-Agent":      ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                        "AppleWebKit/537.36 (KHTML, like Gecko) "
                        "Chrome/124.0.0.0 Safari/537.36"),
    "Accept":          "application/json, text/plain, */*",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer":         "https://www.nseindia.com/",
    "Connection":      "keep-alive",
    "DNT":             "1",
}


def _get_nse_session() -> requests.Session:
    global _NSE_SESSION
    if _NSE_SESSION is not None:
        return _NSE_SESSION
    session = requests.Session()
    session.headers.update(_NSE_HEADERS)
    try:
        session.get("https://www.nseindia.com", timeout=NSE_TIMEOUT)
        time.sleep(0.5)
        session.get("https://www.nseindia.com/api/market-status",
                    timeout=NSE_TIMEOUT)
    except Exception:
        pass
    _NSE_SESSION = session
    return session


def fetch_delivery(symbol: str) -> Optional[float]:
    if not FETCH_DELIVERY:
        return None
    global _NSE_SESSION
    session = _get_nse_session()
    url = f"https://www.nseindia.com/api/quote-equity?symbol={symbol}"
    for attempt in range(2):          # only 2 retries instead of 3
        try:
            resp = session.get(url, timeout=NSE_TIMEOUT)
            if resp.status_code == 401:
                _NSE_SESSION = None
                session = _get_nse_session()
                continue
            if resp.status_code != 200:
                return None
            data = resp.json()
            ti = data.get("tradeInfo", {})
            d_qty = ti.get("deliveryQuantity") or ti.get(
                "delivQty") or ti.get("deliveryQty")
            t_qty = (ti.get("totalTradedQuantity") or ti.get("tradedQuantity")
                     or ti.get("totalTradedQty"))
            if d_qty and t_qty and float(t_qty) > 0:
                return round(float(d_qty) / float(t_qty) * 100, 2)
            dpct = data.get("deliveryToTradedQuantity")
            if dpct is not None:
                return round(float(dpct), 2)
            return None
        except requests.exceptions.ConnectionError:
            return None
        except requests.exceptions.Timeout:
            time.sleep(1.0)
        except Exception:
            return None
    return None


# ═══════════════════════════════════════════════════════════════════════════════
#  6. SIGNAL + WIN/LOSS
# ═══════════════════════════════════════════════════════════════════════════════

def classify_signal(rsi: Optional[float]) -> str:
    if rsi is None or np.isnan(rsi):
        return "N/A"
    if 55 < rsi < 75:
        return "Momentum"
    if rsi < 40:
        return "Weak"
    return "Neutral"


def classify_win_loss(ret_1w: Optional[float]) -> str:
    if ret_1w is None:
        return "N/A"
    return "Win" if ret_1w > 0 else "Loss"


# ═══════════════════════════════════════════════════════════════════════════════
#  7. SINGLE-STOCK PIPELINE
# ═══════════════════════════════════════════════════════════════════════════════

def process_stock(row: pd.Series) -> dict:
    symbol = str(row["Symbol"]).strip().upper()
    sector = row.get("Sector",     "N/A")
    mkt_cap = row.get("Market Cap", "N/A")
    entry_date = pd.to_datetime(row["Date of Entry"])

    base = {
        "Symbol":              symbol,
        "Sector":              sector,
        "Market Cap":          mkt_cap,
        "Original Entry Date": entry_date.date(),
        "Entry Date":          None,
        "Date Shift (Days)":   None,
        "Entry Price":         None,
        "RSI":                 None,
        "1W Return %":         None,
        "2W Return %":         None,
        "3W Return %":         None,
        "Delivery %":          None,
        "Signal":              "N/A",
        "Win/Loss":            "N/A",
    }

    hist_start = entry_date - timedelta(days=200)
    hist_end = entry_date + timedelta(days=35)

    df = fetch_data(symbol, hist_start, hist_end)
    if df is None or df.empty:
        return base

    if "Close" not in df.columns:
        close_candidates = [c for c in df.columns if c.lower() == "close"]
        if not close_candidates:
            return base
        df = df.rename(columns={close_candidates[0]: "Close"})

    entry_idx = get_nearest_trading_day_idx(df, entry_date)
    if entry_idx is None:
        return base

    actual_entry_date = df.index[entry_idx].date()
    shift_days = (actual_entry_date - entry_date.date()).days

    closes = df["Close"].astype(float)
    rsi_series = calculate_rsi(closes, period=14)
    rsi_raw = rsi_series.iloc[entry_idx]
    rsi_val = float(rsi_raw) if not np.isnan(rsi_raw) else None

    returns = compute_returns(df, entry_idx, windows=(5, 10, 15))
    delivery = fetch_delivery(symbol)

    result = {
        **base,
        "Entry Date":        actual_entry_date,
        "Date Shift (Days)": shift_days,
        "Entry Price":       returns["entry_price"],
        "RSI":               round(rsi_val, 2) if rsi_val is not None else None,
        "1W Return %":       returns.get("ret_5d"),
        "2W Return %":       returns.get("ret_10d"),
        "3W Return %":       returns.get("ret_15d"),
        "Delivery %":        delivery,
    }
    result["Signal"] = classify_signal(result["RSI"])
    result["Win/Loss"] = classify_win_loss(result["1W Return %"])

    log.info(
        f"[{symbol}]  Date={actual_entry_date}  Price=₹{result['Entry Price']}  "
        f"RSI={result['RSI']}  1W={result['1W Return %']}%  "
        f"Signal={result['Signal']}  {result['Win/Loss']}"
    )
    return result


# ═══════════════════════════════════════════════════════════════════════════════
#  8. ORCHESTRATION  — parallel with batch pre-fetch
# ═══════════════════════════════════════════════════════════════════════════════

def run_backtest(input_file: str,
                 output_file: str = "final_output.xlsx",
                 progress_callback=None) -> pd.DataFrame:
    """
    Main pipeline:
      1. Load screener file
      2. Batch pre-fetch ALL price data (one yf.download call per chunk)
      3. Process stocks in parallel (ThreadPoolExecutor)
      4. Save professional Excel report

    Parameters
    ----------
    progress_callback : callable(current, total) | None
        If provided, called after each stock completes — use for Streamlit
        progress bars:  progress_callback(i, total)
    """
    df_input = load_input_file(input_file)
    total = len(df_input)

    # ── Step 1: determine global date range for batch download ────────────────
    min_date = df_input["Date of Entry"].min() - timedelta(days=200)
    max_date = df_input["Date of Entry"].max() + timedelta(days=35)

    symbols = df_input["Symbol"].unique().tolist()

    # ── Step 2: batch download in chunks ──────────────────────────────────────
    for i in range(0, len(symbols), BATCH_CHUNK_SIZE):
        chunk = symbols[i: i + BATCH_CHUNK_SIZE]
        _batch_prefetch(chunk, min_date, max_date, exchange="NS")
        # Fallback .BO batch for symbols that got no .NS data
        ns_miss = [
            s for s in chunk
            if not any(
                f"{s}.NS" in k and _DATA_CACHE.get(k) is not None
                for k in _DATA_CACHE
            )
        ]
        if ns_miss:
            _batch_prefetch(ns_miss, min_date, max_date, exchange="BO")

    # ── Step 3: parallel processing ───────────────────────────────────────────
    rows = [None] * total
    futures_map = {}

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        for idx, (_, row) in enumerate(df_input.iterrows()):
            future = executor.submit(process_stock, row)
            futures_map[future] = idx

        completed = 0
        for future in as_completed(futures_map):
            idx = futures_map[future]
            try:
                rows[idx] = future.result()
            except Exception as e:
                row = df_input.iloc[idx]
                log.error(f"[{row['Symbol']}] Unhandled error: {e}")
                rows[idx] = {
                    "Symbol":              row["Symbol"],
                    "Sector":              row.get("Sector",     "N/A"),
                    "Market Cap":          row.get("Market Cap", "N/A"),
                    "Original Entry Date": row["Date of Entry"].date(),
                    "Entry Date":          None,
                    "Date Shift (Days)":   None,
                    "Entry Price":         None,
                    "RSI":                 None,
                    "1W Return %":         None,
                    "2W Return %":         None,
                    "3W Return %":         None,
                    "Delivery %":          None,
                    "Signal":              "N/A",
                    "Win/Loss":            "N/A",
                }
            completed += 1
            if progress_callback:
                progress_callback(completed, total)

    df_out = pd.DataFrame(rows)
    save_excel(df_out, output_file)
    print_summary(df_out)
    return df_out


# ═══════════════════════════════════════════════════════════════════════════════
#  9. EXCEL OUTPUT
# ═══════════════════════════════════════════════════════════════════════════════

C_HEADER_BG = "1F3864"
C_HEADER_FG = "FFFFFF"
C_ALT_ROW = "EEF2F7"
C_MOMENTUM = "C6EFCE"
C_WEAK = "FFC7CE"
C_NEUTRAL = "FFEB9C"
C_WIN = "C6EFCE"
C_LOSS = "FFC7CE"
C_SUMMARY_BG = "2E4057"
C_POS_RET = "375623"
C_NEG_RET = "9C0006"


def _thin_border():
    t = Side(style="thin", color="CCCCCC")
    return Border(left=t, right=t, top=t, bottom=t)


def _fill(color):
    return PatternFill("solid", start_color=color, fgColor=color)


def save_excel(df: pd.DataFrame, filepath: str):
    wb = Workbook()
    _build_main_sheet(wb, df)
    _build_summary_sheet(wb, df)
    wb.save(filepath)
    log.info(f"✅  Saved → {filepath}")


def _build_main_sheet(wb: Workbook, df: pd.DataFrame):
    ws = wb.active
    ws.title = "Backtest Results"
    ws.freeze_panes = "A2"

    columns = [
        ("Symbol",              14),
        ("Sector",              22),
        ("Market Cap",          14),
        ("Original Entry Date", 18),
        ("Entry Date",          14),
        ("Date Shift (Days)",   16),
        ("Entry Price",         14),
        ("RSI",                 10),
        ("1W Return %",         13),
        ("2W Return %",         13),
        ("3W Return %",         13),
        ("Delivery %",          13),
        ("Signal",              13),
        ("Win/Loss",            11),
    ]

    for ci, (hdr, width) in enumerate(columns, 1):
        c = ws.cell(row=1, column=ci, value=hdr)
        c.font = Font(name="Arial", bold=True, color=C_HEADER_FG, size=11)
        c.fill = _fill(C_HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _thin_border()
        ws.column_dimensions[get_column_letter(ci)].width = width

    ws.row_dimensions[1].height = 22
    col_names = [c[0] for c in columns]

    for ri, (_, row) in enumerate(df[col_names].iterrows(), 2):
        is_alt = (ri % 2 == 0)
        for ci, col in enumerate(col_names, 1):
            val = row[col]
            if not isinstance(val, str) and pd.isna(val):
                val = None
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = Font(name="Arial", size=10)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _thin_border()
            if is_alt:
                c.fill = _fill(C_ALT_ROW)

        ws.cell(row=ri, column=4).number_format = "YYYY-MM-DD"

        ep = ws.cell(row=ri, column=5)
        if isinstance(ep.value, (int, float)):
            ep.number_format = "₹#,##0.00"

        rsi_c = ws.cell(row=ri, column=6)
        if isinstance(rsi_c.value, (int, float)):
            rsi_c.number_format = "0.00"

        for ret_col in (9, 10, 11):
            rc = ws.cell(row=ri, column=ret_col)
            rv = rc.value
            if isinstance(rv, (int, float)):
                rc.number_format = "+0.00;-0.00;0.00"
                if rv > 0:
                    rc.font = Font(name="Arial", size=10,
                                   bold=True, color=C_POS_RET)
                elif rv < 0:
                    rc.font = Font(name="Arial", size=10,
                                   bold=True, color=C_NEG_RET)

        dv = ws.cell(row=ri, column=12)
        if isinstance(dv.value, (int, float)):
            dv.number_format = "0.00"

        sc = ws.cell(row=ri, column=13)
        sig = row["Signal"]
        if sig == "Momentum":
            sc.fill = _fill(C_MOMENTUM)
            sc.font = Font(name="Arial", size=10, bold=True, color="375623")
        elif sig == "Weak":
            sc.fill = _fill(C_WEAK)
            sc.font = Font(name="Arial", size=10, bold=True, color="9C0006")
        elif sig == "Neutral":
            sc.fill = _fill(C_NEUTRAL)
            sc.font = Font(name="Arial", size=10, bold=True, color="7D6608")

        wlc = ws.cell(row=ri, column=14)
        if row["Win/Loss"] == "Win":
            wlc.fill = _fill(C_WIN)
            wlc.font = Font(name="Arial", size=10, bold=True, color=C_POS_RET)
        elif row["Win/Loss"] == "Loss":
            wlc.fill = _fill(C_LOSS)
            wlc.font = Font(name="Arial", size=10, bold=True, color=C_NEG_RET)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"


def _build_summary_sheet(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("Summary")
    valid = df[df["Entry Price"].notna()]
    total = len(df)

    win_df = valid[valid["Win/Loss"] == "Win"]
    loss_df = valid[valid["Win/Loss"] == "Loss"]
    win_rate = (len(win_df) / len(valid) * 100) if len(valid) else 0
    avg_1w = valid["1W Return %"].mean()
    avg_2w = valid["2W Return %"].mean()
    avg_3w = valid["3W Return %"].mean()
    sig_counts = valid["Signal"].value_counts().to_dict()

    def section(row, title):
        c = ws.cell(row=row, column=2, value=title)
        c.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
        c.fill = _fill("2E4057")
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(f"B{row}:F{row}")
        c.border = _thin_border()

    def metric(row, label, value, fmt="general"):
        lc = ws.cell(row=row, column=2, value=label)
        vc = ws.cell(row=row, column=4, value=value)
        lc.font = Font(name="Arial", bold=True, size=11)
        vc.font = Font(name="Arial", size=11, color="1F3864")
        vc.alignment = Alignment(horizontal="center")
        lc.alignment = Alignment(vertical="center")
        ws.merge_cells(f"B{row}:C{row}")
        ws.merge_cells(f"D{row}:F{row}")
        lc.border = _thin_border()
        vc.border = _thin_border()
        if fmt == "pct":
            vc.number_format = "0.00%"
        elif fmt == "ret":
            vc.number_format = "+0.00%;-0.00%;0.00%"
        elif fmt == "currency":
            vc.number_format = "₹#,##0.00"

    title_c = ws["B2"]
    title_c.value = "📊  BACKTEST SUMMARY REPORT"
    title_c.font = Font(name="Arial", bold=True, size=16, color=C_HEADER_FG)
    title_c.fill = _fill(C_HEADER_BG)
    title_c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("B2:F2")
    ws.row_dimensions[2].height = 32

    r = 4
    section(r, "  OVERVIEW")
    r += 1
    metric(r,  "Total Stocks Processed",  total)
    r += 1
    metric(r,  "Successfully Fetched",    len(valid))
    r += 1
    metric(r,  "Stocks Skipped / Error",  total - len(valid))
    r += 2

    section(r, "  PERFORMANCE METRICS")
    r += 1
    metric(r,  "Win Rate (1W)",           win_rate / 100,      "pct")
    r += 1
    metric(r,  "Total Wins",              len(win_df))
    r += 1
    metric(r,  "Total Losses",            len(loss_df))
    r += 1
    metric(r,  "Avg 1-Week Return",       (avg_1w or 0) / 100, "ret")
    r += 1
    metric(r,  "Avg 2-Week Return",       (avg_2w or 0) / 100, "ret")
    r += 1
    metric(r,  "Avg 3-Week Return",       (avg_3w or 0) / 100, "ret")
    r += 2

    section(r, "  SIGNAL BREAKDOWN")
    r += 1
    metric(r,  "Momentum Stocks", sig_counts.get("Momentum", 0))
    r += 1
    metric(r,  "Neutral Stocks",  sig_counts.get("Neutral",  0))
    r += 1
    metric(r,  "Weak Stocks",     sig_counts.get("Weak",     0))
    r += 2

    section(r, "  SECTOR DISTRIBUTION")
    r += 1
    sec_grp = valid.groupby("Sector").size().sort_values(ascending=False)
    for sector_name, cnt in sec_grp.head(12).items():
        metric(r, f"  {sector_name}", cnt)
        r += 1

    for col, width in [("A", 2), ("B", 28), ("C", 5),
                       ("D", 18), ("E", 5), ("F", 5)]:
        ws.column_dimensions[col].width = width


# ═══════════════════════════════════════════════════════════════════════════════
#  10. CONSOLE SUMMARY
# ═══════════════════════════════════════════════════════════════════════════════

def print_summary(df: pd.DataFrame):
    valid = df[df["Entry Price"].notna()]
    total = len(df)
    wins = (valid["Win/Loss"] == "Win").sum()
    win_rate = wins / len(valid) * 100 if len(valid) else 0

    print("\n" + "═" * 58)
    print("  📊  BACKTEST SUMMARY")
    print("═" * 58)
    print(f"  Total stocks           : {total}")
    print(f"  Successfully analysed  : {len(valid)}")
    print(f"  Skipped / errors       : {total - len(valid)}")
    print(f"  Win rate (1W)          : {win_rate:.1f}%")
    if len(valid):
        print(f"  Avg 1W Return          : {valid['1W Return %'].mean():.2f}%")
        print(f"  Avg 2W Return          : {valid['2W Return %'].mean():.2f}%")
        print(f"  Avg 3W Return          : {valid['3W Return %'].mean():.2f}%")
        sig_counts = valid["Signal"].value_counts()
        print(f"\n  Signal breakdown:")
        for sig, cnt in sig_counts.items():
            print(f"    {sig:12s} : {cnt}")
    print("═" * 58 + "\n")


# ═══════════════════════════════════════════════════════════════════════════════
#  STREAMLIT HELPER  — drop this into your app.py
# ═══════════════════════════════════════════════════════════════════════════════
#
#  import streamlit as st
#  from backtest import run_backtest
#
#  uploaded = st.file_uploader("Upload screener file", type=["xlsx","csv"])
#  if uploaded and st.button("Run Backtest"):
#      with st.spinner("Running backtest …"):
#          progress_bar = st.progress(0)
#          def on_progress(done, total):
#              progress_bar.progress(done / total)
#          df = run_backtest(uploaded.name, "final_output.xlsx",
#                            progress_callback=on_progress)
#      st.success("Done!")
#      st.dataframe(df)
#      with open("final_output.xlsx", "rb") as f:
#          st.download_button("Download Excel", f, "backtest_results.xlsx")


# ═══════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    import sys

    input_path = sys.argv[1] if len(sys.argv) > 1 else "screener_input.xlsx"
    output_path = sys.argv[2] if len(sys.argv) > 2 else "final_output.xlsx"

    if not os.path.exists(input_path):
        print(f"❌  Input file not found: '{input_path}'")
        print("Usage: python backtest.py <input_file> [output_file]")
        sys.exit(1)

    log.info(f"Input  : {input_path}")
    log.info(f"Output : {output_path}")
    run_backtest(input_path, output_path)
